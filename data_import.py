"""
Módulo de Importación de Datos
Permite importar datos masivos desde archivos CSV/Excel
"""

import csv
import logging
from pathlib import Path
from typing import List, Dict, Tuple
import openpyxl
from database import SistemaDAO

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class DataImporter:
    def __init__(self, dao: SistemaDAO):
        """
        Inicializa el importador de datos
        
        Args:
            dao: Instancia de SistemaDAO para acceso a la base de datos
        """
        self.dao = dao
        self.errores = []
        self.advertencias = []
    
    def validar_docente(self, row: dict) -> Tuple[bool, str]:
        """Valida los datos de un docente"""
        if not row.get('nombre'):
            return False, "Nombre es requerido"
        
        if row.get('email'):
            # Verificar formato básico de email
            if '@' not in row['email']:
                return False, f"Email inválido: {row['email']}"
        
        if row.get('horas_contratadas'):
            try:
                horas = int(row['horas_contratadas'])
                if horas < 0 or horas > 44:
                    return False, f"Horas contratadas fuera de rango (0-44): {horas}"
            except ValueError:
                return False, f"Horas contratadas debe ser un número: {row['horas_contratadas']}"
        
        if row.get('evaluacion'):
            try:
                eval_val = float(row['evaluacion'])
                if eval_val < 0 or eval_val > 5:
                    return False, f"Evaluación fuera de rango (0-5): {eval_val}"
            except ValueError:
                return False, f"Evaluación debe ser un número: {row['evaluacion']}"
        
        return True, ""
    
    def validar_sala(self, row: dict) -> Tuple[bool, str]:
        """Valida los datos de una sala"""
        if not row.get('nombre'):
            return False, "Nombre es requerido"
        
        if row.get('capacidad'):
            try:
                cap = int(row['capacidad'])
                if cap < 0:
                    return False, f"Capacidad no puede ser negativa: {cap}"
            except ValueError:
                return False, f"Capacidad debe ser un número: {row['capacidad']}"
        
        return True, ""
    
    def validar_carrera(self, row: dict) -> Tuple[bool, str]:
        """Valida los datos de una carrera"""
        if not row.get('nombre'):
            return False, "Nombre es requerido"
        
        if row.get('alumnos_proyectados'):
            try:
                alumnos = int(row['alumnos_proyectados'])
                if alumnos < 0:
                    return False, f"Alumnos proyectados no puede ser negativo: {alumnos}"
            except ValueError:
                return False, f"Alumnos proyectados debe ser un número: {row['alumnos_proyectados']}"
        
        return True, ""
    
    def validar_modulo(self, row: dict) -> Tuple[bool, str]:
        """Valida los datos de un módulo"""
        if not row.get('nombre'):
            return False, "Nombre es requerido"
        
        if not row.get('carrera_id'):
            return False, "Carrera ID es requerido"
        
        # Validar que la carrera existe
        try:
            carrera_id = int(row['carrera_id'])
            carreras = self.dao.obtener_carreras()
            if not any(c['id'] == carrera_id for c in carreras):
                return False, f"Carrera ID {carrera_id} no existe"
        except ValueError:
            return False, f"Carrera ID debe ser un número: {row['carrera_id']}"
        
        return True, ""
    
    def importar_docentes_csv(self, filepath: str, limpiar_existentes: bool = False) -> int:
        """
        Importa docentes desde un archivo CSV
        
        Args:
            filepath: Ruta al archivo CSV
            limpiar_existentes: Si True, elimina todos los docentes existentes antes de importar
        
        Returns:
            int: Número de docentes importados
        """
        self.errores = []
        self.advertencias = []
        
        if limpiar_existentes:
            logging.warning("⚠️ Limpiando docentes existentes...")
            # Nota: Esto podría fallar si hay módulos asignados
        
        importados = 0
        
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            for i, row in enumerate(reader, start=2):  # Empezar en 2 por el encabezado
                # Validar
                valido, error = self.validar_docente(row)
                if not valido:
                    self.errores.append(f"Fila {i}: {error}")
                    continue
                
                # Preparar datos
                docente_data = (
                    row['nombre'],
                    row.get('titulo', ''),
                    row.get('contrato', 'Honorarios'),
                    int(row.get('horas_contratadas', 0)),
                    row.get('email', ''),
                    float(row.get('evaluacion', 0.0))
                )
                
                # Insertar
                try:
                    if row.get('id'):
                        # Actualizar existente
                        docente_id = int(row['id'])
                        self.dao.guardar_docente(docente_data, docente_id)
                    else:
                        # Crear nuevo
                        self.dao.guardar_docente(docente_data)
                    
                    importados += 1
                except Exception as e:
                    self.errores.append(f"Fila {i}: Error al guardar - {str(e)}")
        
        logging.info(f"✅ Importados {importados} docentes")
        if self.errores:
            logging.warning(f"⚠️ {len(self.errores)} errores encontrados")
        
        return importados
    
    def importar_salas_csv(self, filepath: str, limpiar_existentes: bool = False) -> int:
        """Importa salas desde un archivo CSV"""
        self.errores = []
        importados = 0
        
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            for i, row in enumerate(reader, start=2):
                valido, error = self.validar_sala(row)
                if not valido:
                    self.errores.append(f"Fila {i}: {error}")
                    continue
                
                sala_data = (
                    row['nombre'],
                    int(row.get('capacidad', 0)),
                    row.get('tipo', 'Teórica')
                )
                
                try:
                    if row.get('id'):
                        sala_id = int(row['id'])
                        self.dao.guardar_sala(sala_data, sala_id)
                    else:
                        self.dao.guardar_sala(sala_data)
                    
                    importados += 1
                except Exception as e:
                    self.errores.append(f"Fila {i}: Error al guardar - {str(e)}")
        
        logging.info(f"✅ Importadas {importados} salas")
        return importados
    
    def importar_carreras_csv(self, filepath: str, limpiar_existentes: bool = False) -> int:
        """Importa carreras desde un archivo CSV"""
        self.errores = []
        importados = 0
        
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            for i, row in enumerate(reader, start=2):
                valido, error = self.validar_carrera(row)
                if not valido:
                    self.errores.append(f"Fila {i}: {error}")
                    continue
                
                carrera_data = (
                    row['nombre'],
                    row.get('jornada', 'Diurna'),
                    int(row.get('alumnos_proyectados', 0))
                )
                
                # Parsear semestres
                semestres_str = row.get('semestres', '')
                semestres = []
                if semestres_str:
                    try:
                        semestres = [int(s.strip()) for s in semestres_str.split(',') if s.strip()]
                    except ValueError:
                        self.errores.append(f"Fila {i}: Formato de semestres inválido: {semestres_str}")
                        continue
                
                # Parsear salas
                salas_str = row.get('salas_ids', '')
                salas_ids = []
                if salas_str:
                    try:
                        salas_ids = [int(s.strip()) for s in salas_str.split(',') if s.strip()]
                    except ValueError:
                        self.errores.append(f"Fila {i}: Formato de salas inválido: {salas_str}")
                        continue
                
                try:
                    if row.get('id'):
                        carrera_id = int(row['id'])
                        self.dao.guardar_carrera(carrera_data, semestres, salas_ids, carrera_id)
                    else:
                        self.dao.guardar_carrera(carrera_data, semestres, salas_ids)
                    
                    importados += 1
                except Exception as e:
                    self.errores.append(f"Fila {i}: Error al guardar - {str(e)}")
        
        logging.info(f"✅ Importadas {importados} carreras")
        return importados
    
    def importar_modulos_csv(self, filepath: str, limpiar_existentes: bool = False) -> int:
        """Importa módulos desde un archivo CSV"""
        self.errores = []
        importados = 0
        
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            for i, row in enumerate(reader, start=2):
                valido, error = self.validar_modulo(row)
                if not valido:
                    self.errores.append(f"Fila {i}: {error}")
                    continue
                
                modulo_data = (
                    row['nombre'],
                    row.get('codigo', ''),
                    int(row.get('horas_teoricas', 0)),
                    int(row.get('horas_practicas', 0)),
                    int(row.get('alumnos_proyectados', 0)),
                    int(row['carrera_id']),
                    int(row.get('semestre', 1)),
                    int(row['docente_id']) if row.get('docente_id') else None,
                    int(row['sala_id']) if row.get('sala_id') else None
                )
                
                try:
                    if row.get('id'):
                        modulo_id = int(row['id'])
                        self.dao.guardar_modulo(modulo_data, [], modulo_id)
                    else:
                        self.dao.guardar_modulo(modulo_data, [])
                    
                    importados += 1
                except Exception as e:
                    self.errores.append(f"Fila {i}: Error al guardar - {str(e)}")
        
        logging.info(f"✅ Importados {importados} módulos")
        return importados
    
    def importar_horarios_modulos_csv(self, filepath: str) -> int:
        """Importa horarios de módulos desde un archivo CSV"""
        self.errores = []
        importados = 0
        
        # Agrupar horarios por módulo
        horarios_por_modulo = {}
        
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            
            for i, row in enumerate(reader, start=2):
                try:
                    modulo_id = int(row['modulo_id'])
                    
                    if modulo_id not in horarios_por_modulo:
                        horarios_por_modulo[modulo_id] = []
                    
                    horarios_por_modulo[modulo_id].append({
                        'dia': row['dia'],
                        'hora_inicio': row['hora_inicio'],
                        'hora_fin': row['hora_fin']
                    })
                except Exception as e:
                    self.errores.append(f"Fila {i}: Error - {str(e)}")
        
        # Guardar horarios
        for modulo_id, horarios in horarios_por_modulo.items():
            try:
                # Obtener datos del módulo
                modulos = self.dao.obtener_modulos()
                modulo = next((m for m in modulos if m['id'] == modulo_id), None)
                
                if not modulo:
                    self.errores.append(f"Módulo ID {modulo_id} no existe")
                    continue
                
                # Preparar datos del módulo
                modulo_data = (
                    modulo['nombre'],
                    modulo.get('codigo', ''),
                    modulo.get('horas_teoricas', 0),
                    modulo.get('horas_practicas', 0),
                    modulo.get('alumnos_proyectados', 0),
                    modulo.get('carrera_id'),
                    modulo.get('semestre', 1),
                    modulo.get('docente_id'),
                    modulo.get('sala_id')
                )
                
                # Guardar con horarios
                self.dao.guardar_modulo(modulo_data, horarios, modulo_id)
                importados += len(horarios)
                
            except Exception as e:
                self.errores.append(f"Módulo {modulo_id}: Error al guardar horarios - {str(e)}")
        
        logging.info(f"✅ Importados {importados} horarios")
        return importados
    
    def importar_desde_excel(self, filepath: str, limpiar_existentes: bool = False) -> dict:
        """
        Importa datos desde un archivo Excel con múltiples hojas
        
        Args:
            filepath: Ruta al archivo Excel
            limpiar_existentes: Si True, limpia datos existentes antes de importar
        
        Returns:
            dict: Resumen de importación
        """
        logging.info(f"📊 Importando desde Excel: {filepath}")
        
        wb = openpyxl.load_workbook(filepath)
        resultados = {}
        
        # Importar en orden de dependencias
        orden = ['Salas', 'Docentes', 'Carreras', 'Módulos', 'Horarios Módulos']
        
        for sheet_name in orden:
            if sheet_name not in wb.sheetnames:
                logging.warning(f"⚠️ Hoja '{sheet_name}' no encontrada")
                continue
            
            ws = wb[sheet_name]
            
            # Convertir a CSV temporal
            temp_csv = Path(f"temp_{sheet_name.lower().replace(' ', '_')}.csv")
            
            with open(temp_csv, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            
            # Importar según el tipo
            try:
                if sheet_name == 'Docentes':
                    count = self.importar_docentes_csv(str(temp_csv), limpiar_existentes)
                    resultados['docentes'] = count
                elif sheet_name == 'Salas':
                    count = self.importar_salas_csv(str(temp_csv), limpiar_existentes)
                    resultados['salas'] = count
                elif sheet_name == 'Carreras':
                    count = self.importar_carreras_csv(str(temp_csv), limpiar_existentes)
                    resultados['carreras'] = count
                elif sheet_name == 'Módulos':
                    count = self.importar_modulos_csv(str(temp_csv), limpiar_existentes)
                    resultados['modulos'] = count
                elif sheet_name == 'Horarios Módulos':
                    count = self.importar_horarios_modulos_csv(str(temp_csv))
                    resultados['horarios'] = count
            except Exception as e:
                logging.error(f"❌ Error importando {sheet_name}: {e}")
                resultados[sheet_name] = 0
            finally:
                # Limpiar archivo temporal
                if temp_csv.exists():
                    temp_csv.unlink()
        
        logging.info("✅ Importación desde Excel completada")
        return resultados
    
    def obtener_reporte_errores(self) -> str:
        """Genera un reporte de errores de la última importación"""
        if not self.errores:
            return "✅ No se encontraron errores"
        
        reporte = f"⚠️ Se encontraron {len(self.errores)} errores:\n\n"
        for error in self.errores[:50]:  # Limitar a 50 errores
            reporte += f"• {error}\n"
        
        if len(self.errores) > 50:
            reporte += f"\n... y {len(self.errores) - 50} errores más"
        
        return reporte


# Función de utilidad para uso directo
def importar_datos_desde_excel(filepath: str, limpiar_existentes: bool = False):
    """
    Importa datos desde un archivo Excel
    
    Args:
        filepath: Ruta al archivo Excel
        limpiar_existentes: Si True, limpia datos existentes
    
    Returns:
        dict: Resumen de importación
    """
    dao = SistemaDAO()
    importer = DataImporter(dao)
    
    resultados = importer.importar_desde_excel(filepath, limpiar_existentes)
    
    print("\n📊 Resumen de Importación:")
    for tabla, count in resultados.items():
        print(f"   - {tabla}: {count} registros")
    
    if importer.errores:
        print(f"\n{importer.obtener_reporte_errores()}")
    
    return resultados


if __name__ == "__main__":
    # Prueba de importación
    import sys
    
    if len(sys.argv) > 1:
        archivo = sys.argv[1]
        print(f"🚀 Importando datos desde: {archivo}")
        importar_datos_desde_excel(archivo)
    else:
        print("Uso: python data_import.py <archivo.xlsx>")
