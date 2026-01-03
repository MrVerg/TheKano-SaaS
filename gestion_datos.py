"""
Gestión de Datos - Importación y Exportación
Interfaz de usuario para importar y exportar datos masivos
"""

import flet as ft
import logging
from pathlib import Path
from data_export import DataExporter
from data_import import DataImporter
from database import SistemaDAO

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class GestionDatos:
    def __init__(self, page: ft.Page, dao: SistemaDAO):
        """
        Inicializa la vista de gestión de datos
        
        Args:
            page: Página de Flet
            dao: Instancia de SistemaDAO
        """
        self.page = page
        self.dao = dao
        self.exporter = DataExporter(dao)
        self.importer = DataImporter(dao)
        
        # Componentes UI
        self.progress_bar = None
        self.status_text = None
        self.resultado_text = None
    
    def crear_vista(self) -> ft.Container:
        """Crea la vista de gestión de datos"""
        
        # Título
        titulo = ft.Text(
            "Gestión de Datos",
            size=28,
            weight=ft.FontWeight.BOLD,
            color="#1a237e"
        )
        
        # Sección de Exportación
        export_title = ft.Text(
            "📤 Exportar Datos",
            size=20,
            weight=ft.FontWeight.BOLD,
            color="#2196F3"
        )
        
        export_desc = ft.Text(
            "Exporta todos los datos de la base de datos a formato Excel o CSV",
            size=14,
            color="#666"
        )
        
        btn_export_excel = ft.ElevatedButton(
            "Exportar a Excel",
            icon=ft.Icons.TABLE_CHART,
            on_click=self.exportar_excel,
            bgcolor="#4CAF50",
            color="white",
            height=50,
            width=200
        )
        
        btn_export_csv = ft.ElevatedButton(
            "Exportar a CSV",
            icon=ft.Icons.FOLDER_ZIP,
            on_click=self.exportar_csv,
            bgcolor="#2196F3",
            color="white",
            height=50,
            width=200
        )
        
        export_buttons = ft.Row(
            [btn_export_excel, btn_export_csv],
            spacing=20
        )
        
        # Sección de Importación
        import_title = ft.Text(
            "📥 Importar Datos",
            size=20,
            weight=ft.FontWeight.BOLD,
            color="#FF9800"
        )
        
        import_desc = ft.Text(
            "Importa datos masivos desde archivos Excel o CSV",
            size=14,
            color="#666"
        )
        
        btn_import_excel = ft.ElevatedButton(
            "Importar desde Excel",
            icon=ft.Icons.UPLOAD_FILE,
            on_click=self.importar_excel,
            bgcolor="#FF9800",
            color="white",
            height=50,
            width=200
        )
        
        btn_import_csv = ft.ElevatedButton(
            "Importar desde CSV",
            icon=ft.Icons.FOLDER_OPEN,
            on_click=self.importar_csv_folder,
            bgcolor="#FF5722",
            color="white",
            height=50,
            width=200
        )
        
        import_buttons = ft.Row(
            [btn_import_excel, btn_import_csv],
            spacing=20
        )
        
        # Barra de progreso
        self.progress_bar = ft.ProgressBar(
            width=600,
            visible=False,
            color="#4CAF50"
        )
        
        # Texto de estado
        self.status_text = ft.Text(
            "",
            size=14,
            color="#666",
            visible=False
        )
        
        # Resultado
        self.resultado_text = ft.Container(
            content=ft.Text("", size=14),
            bgcolor="#f5f5f5",
            padding=15,
            border_radius=8,
            visible=False
        )
        
        # Información adicional
        info_card = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.INFO_OUTLINE, color="#2196F3", size=24),
                    ft.Text("Información Importante", size=16, weight=ft.FontWeight.BOLD)
                ]),
                ft.Divider(height=1, color="#e0e0e0"),
                ft.Text("• Los archivos exportados se guardan en la carpeta 'exports'", size=13),
                ft.Text("• Al importar, los datos existentes se actualizan o se crean nuevos", size=13),
                ft.Text("• Formato Excel: un archivo con múltiples hojas", size=13),
                ft.Text("• Formato CSV: múltiples archivos, uno por tabla", size=13),
                ft.Text("• Se validan los datos antes de importar", size=13),
            ], spacing=8),
            bgcolor="#E3F2FD",
            padding=20,
            border_radius=10,
            border=ft.border.all(1, "#2196F3")
        )
        
        # Plantilla de ejemplo
        plantilla_card = ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.DOWNLOAD, color="#4CAF50", size=24),
                    ft.Text("Plantilla de Importación", size=16, weight=ft.FontWeight.BOLD)
                ]),
                ft.Divider(height=1, color="#e0e0e0"),
                ft.Text("Descarga una plantilla vacía para importar datos:", size=13),
                ft.ElevatedButton(
                    "Descargar Plantilla Excel",
                    icon=ft.Icons.DOWNLOAD,
                    on_click=self.descargar_plantilla,
                    bgcolor="#4CAF50",
                    color="white"
                )
            ], spacing=10),
            bgcolor="#E8F5E9",
            padding=20,
            border_radius=10,
            border=ft.border.all(1, "#4CAF50")
        )
        
        # Layout principal
        contenido = ft.Column([
            titulo,
            ft.Divider(height=20, color="transparent"),
            
            # Exportación
            export_title,
            export_desc,
            ft.Divider(height=10, color="transparent"),
            export_buttons,
            ft.Divider(height=30, color="transparent"),
            
            # Importación
            import_title,
            import_desc,
            ft.Divider(height=10, color="transparent"),
            import_buttons,
            ft.Divider(height=30, color="transparent"),
            
            # Progreso y estado
            self.progress_bar,
            self.status_text,
            self.resultado_text,
            ft.Divider(height=20, color="transparent"),
            
            # Información
            info_card,
            ft.Divider(height=15, color="transparent"),
            plantilla_card,
            
        ], spacing=10, scroll=ft.ScrollMode.AUTO)
        
        return ft.Container(
            content=contenido,
            padding=30,
            expand=True
        )
    
    def mostrar_progreso(self, mensaje: str):
        """Muestra la barra de progreso con un mensaje"""
        self.progress_bar.visible = True
        self.status_text.value = mensaje
        self.status_text.visible = True
        self.resultado_text.visible = False
        self.page.update()
    
    def ocultar_progreso(self):
        """Oculta la barra de progreso"""
        self.progress_bar.visible = False
        self.status_text.visible = False
        self.page.update()
    
    def mostrar_resultado(self, mensaje: str, es_error: bool = False):
        """Muestra el resultado de una operación"""
        self.resultado_text.content.value = mensaje
        self.resultado_text.bgcolor = "#FFEBEE" if es_error else "#E8F5E9"
        self.resultado_text.visible = True
        self.ocultar_progreso()
        self.page.update()
    
    def exportar_excel(self, e):
        """Exporta todos los datos a Excel"""
        try:
            self.mostrar_progreso("📊 Exportando datos a Excel...")
            
            filepath = self.exporter.exportar_todo_excel()
            
            mensaje = f"✅ Exportación completada exitosamente\n\n"
            mensaje += f"📄 Archivo generado:\n{filepath}\n\n"
            mensaje += "El archivo contiene las siguientes hojas:\n"
            mensaje += "• Docentes\n• Salas\n• Carreras\n• Módulos\n• Horarios Módulos\n• Disponibilidad Docentes"
            
            self.mostrar_resultado(mensaje)
            
            # Abrir carpeta de exportación
            import os
            os.startfile(Path(filepath).parent)
            
        except Exception as ex:
            logging.error(f"Error exportando a Excel: {ex}")
            self.mostrar_resultado(f"❌ Error al exportar:\n{str(ex)}", es_error=True)
    
    def exportar_csv(self, e):
        """Exporta todos los datos a CSV"""
        try:
            self.mostrar_progreso("📁 Exportando datos a CSV...")
            
            archivos = self.exporter.exportar_todo_csv()
            
            mensaje = f"✅ Exportación completada exitosamente\n\n"
            mensaje += f"📁 Archivos generados ({len(archivos)}):\n\n"
            for nombre, path in archivos.items():
                mensaje += f"• {nombre}: {Path(path).name}\n"
            
            mensaje += f"\n📂 Ubicación: {Path(list(archivos.values())[0]).parent}"
            
            self.mostrar_resultado(mensaje)
            
            # Abrir carpeta de exportación
            import os
            os.startfile(Path(list(archivos.values())[0]).parent)
            
        except Exception as ex:
            logging.error(f"Error exportando a CSV: {ex}")
            self.mostrar_resultado(f"❌ Error al exportar:\n{str(ex)}", es_error=True)
    
    def importar_excel(self, e):
        """Importa datos desde un archivo Excel"""
        def on_file_picked(e: ft.FilePickerResultEvent):
            if e.files:
                filepath = e.files[0].path
                try:
                    self.mostrar_progreso(f"📥 Importando datos desde Excel...")
                    
                    resultados = self.importer.importar_desde_excel(filepath, limpiar_existentes=False)
                    
                    mensaje = "✅ Importación completada\n\n"
                    mensaje += "📊 Resumen:\n"
                    for tabla, count in resultados.items():
                        mensaje += f"• {tabla}: {count} registros\n"
                    
                    if self.importer.errores:
                        mensaje += f"\n⚠️ Se encontraron {len(self.importer.errores)} errores\n"
                        mensaje += "\nPrimeros errores:\n"
                        for error in self.importer.errores[:5]:
                            mensaje += f"• {error}\n"
                    
                    self.mostrar_resultado(mensaje, es_error=bool(self.importer.errores))
                    
                except Exception as ex:
                    logging.error(f"Error importando Excel: {ex}")
                    self.mostrar_resultado(f"❌ Error al importar:\n{str(ex)}", es_error=True)
        
        file_picker = ft.FilePicker(on_result=on_file_picked)
        self.page.overlay.append(file_picker)
        self.page.update()
        
        file_picker.pick_files(
            dialog_title="Seleccionar archivo Excel",
            allowed_extensions=["xlsx", "xls"],
            allow_multiple=False
        )
    
    def importar_csv_folder(self, e):
        """Importa datos desde una carpeta con archivos CSV"""
        def on_folder_picked(e: ft.FilePickerResultEvent):
            if e.path:
                folder_path = Path(e.path)
                try:
                    self.mostrar_progreso(f"📥 Importando datos desde CSV...")
                    
                    total_importados = 0
                    resultados = {}
                    
                    # Buscar archivos CSV en la carpeta
                    archivos_csv = {
                        'docentes': folder_path / 'docentes.csv',
                        'salas': folder_path / 'salas.csv',
                        'carreras': folder_path / 'carreras.csv',
                        'modulos': folder_path / 'modulos.csv',
                        'horarios': folder_path / 'modulo_horarios.csv'
                    }
                    
                    for nombre, archivo in archivos_csv.items():
                        if archivo.exists():
                            if nombre == 'docentes':
                                count = self.importer.importar_docentes_csv(str(archivo))
                            elif nombre == 'salas':
                                count = self.importer.importar_salas_csv(str(archivo))
                            elif nombre == 'carreras':
                                count = self.importer.importar_carreras_csv(str(archivo))
                            elif nombre == 'modulos':
                                count = self.importer.importar_modulos_csv(str(archivo))
                            elif nombre == 'horarios':
                                count = self.importer.importar_horarios_modulos_csv(str(archivo))
                            
                            resultados[nombre] = count
                            total_importados += count
                    
                    mensaje = "✅ Importación completada\n\n"
                    mensaje += "📊 Resumen:\n"
                    for tabla, count in resultados.items():
                        mensaje += f"• {tabla}: {count} registros\n"
                    
                    mensaje += f"\nTotal: {total_importados} registros importados"
                    
                    if self.importer.errores:
                        mensaje += f"\n\n⚠️ Se encontraron {len(self.importer.errores)} errores"
                    
                    self.mostrar_resultado(mensaje, es_error=bool(self.importer.errores))
                    
                except Exception as ex:
                    logging.error(f"Error importando CSV: {ex}")
                    self.mostrar_resultado(f"❌ Error al importar:\n{str(ex)}", es_error=True)
        
        file_picker = ft.FilePicker(on_result=on_folder_picked)
        self.page.overlay.append(file_picker)
        self.page.update()
        
        file_picker.get_directory_path(dialog_title="Seleccionar carpeta con archivos CSV")
    
    def descargar_plantilla(self, e):
        """Genera una plantilla vacía para importación"""
        try:
            self.mostrar_progreso("📝 Generando plantilla...")
            
            # Crear un Excel vacío con las estructuras correctas
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Hoja de Docentes
            ws = wb.create_sheet("Docentes")
            ws.append(['id', 'nombre', 'titulo', 'contrato', 'horas_contratadas', 'email', 'evaluacion'])
            ws.append(['', 'Juan Pérez', 'Ingeniero Civil', 'Planta', '44', 'juan@ceduc.cl', '4.5'])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Hoja de Salas
            ws = wb.create_sheet("Salas")
            ws.append(['id', 'nombre', 'capacidad', 'tipo'])
            ws.append(['', 'Sala A101', '30', 'Teórica'])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Hoja de Carreras
            ws = wb.create_sheet("Carreras")
            ws.append(['id', 'nombre', 'jornada', 'alumnos_proyectados', 'semestres', 'salas_ids'])
            ws.append(['', 'Ingeniería en Informática', 'Diurna', '120', '1,2,3,4,5,6,7,8', '1,2,3'])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Hoja de Módulos
            ws = wb.create_sheet("Módulos")
            ws.append(['id', 'nombre', 'codigo', 'horas_teoricas', 'horas_practicas', 
                      'alumnos_proyectados', 'semestre', 'carrera_id', 'docente_id', 'sala_id'])
            ws.append(['', 'Programación I', 'INF101', '4', '2', '30', '1', '1', '1', '1'])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Hoja de Horarios
            ws = wb.create_sheet("Horarios Módulos")
            ws.append(['modulo_id', 'dia', 'hora_inicio', 'hora_fin'])
            ws.append(['1', 'LUNES', '08:30', '10:15'])
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Guardar
            filepath = Path("exports") / "plantilla_importacion.xlsx"
            filepath.parent.mkdir(exist_ok=True)
            wb.save(filepath)
            
            mensaje = f"✅ Plantilla generada exitosamente\n\n"
            mensaje += f"📄 Archivo: {filepath}\n\n"
            mensaje += "Instrucciones:\n"
            mensaje += "1. Abre el archivo en Excel\n"
            mensaje += "2. Completa los datos en cada hoja\n"
            mensaje += "3. Elimina las filas de ejemplo\n"
            mensaje += "4. Guarda el archivo\n"
            mensaje += "5. Importa usando el botón 'Importar desde Excel'"
            
            self.mostrar_resultado(mensaje)
            
            # Abrir carpeta
            import os
            os.startfile(filepath.parent)
            
        except Exception as ex:
            logging.error(f"Error generando plantilla: {ex}")
            self.mostrar_resultado(f"❌ Error al generar plantilla:\n{str(ex)}", es_error=True)


def main(page: ft.Page):
    """Función principal para prueba standalone"""
    page.title = "Gestión de Datos"
    page.window_width = 900
    page.window_height = 800
    
    dao = SistemaDAO()
    gestion = GestionDatos(page, dao)
    
    page.add(gestion.crear_vista())


if __name__ == "__main__":
    ft.app(target=main)
