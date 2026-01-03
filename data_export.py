"""
Módulo de Exportación de Datos
Permite exportar todos los datos de la base de datos a formato CSV/Excel
"""

import csv
import logging
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from database import SistemaDAO

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class DataExporter:
    def __init__(self, dao: SistemaDAO):
        """
        Inicializa el exportador de datos
        
        Args:
            dao: Instancia de SistemaDAO para acceso a la base de datos
        """
        self.dao = dao
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)
    
    def exportar_docentes_csv(self, filepath: str = None) -> str:
        """Exporta todos los docentes a un archivo CSV"""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"docentes_{timestamp}.csv"
        
        docentes = self.dao.obtener_docentes()
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Encabezados
            writer.writerow(['id', 'nombre', 'titulo', 'contrato', 'horas_contratadas', 'email', 'evaluacion'])
            
            # Datos
            for d in docentes:
                writer.writerow([
                    d['id'],
                    d['nombre'],
                    d.get('titulo', ''),
                    d.get('contrato', ''),
                    d.get('horas_contratadas', 0),
                    d.get('email', ''),
                    d.get('evaluacion', 0.0)
                ])
        
        logging.info(f"✅ Exportados {len(docentes)} docentes a {filepath}")
        return str(filepath)
    
    def exportar_salas_csv(self, filepath: str = None) -> str:
        """Exporta todas las salas a un archivo CSV"""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"salas_{timestamp}.csv"
        
        salas = self.dao.obtener_salas()
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Encabezados
            writer.writerow(['id', 'nombre', 'capacidad', 'tipo'])
            
            # Datos
            for s in salas:
                writer.writerow([
                    s['id'],
                    s['nombre'],
                    s.get('capacidad', 0),
                    s.get('tipo', '')
                ])
        
        logging.info(f"✅ Exportadas {len(salas)} salas a {filepath}")
        return str(filepath)
    
    def exportar_carreras_csv(self, filepath: str = None) -> str:
        """Exporta todas las carreras a un archivo CSV con semestres y salas"""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"carreras_{timestamp}.csv"
        
        carreras = self.dao.obtener_carreras()
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Encabezados
            writer.writerow(['id', 'nombre', 'jornada', 'alumnos_proyectados', 'semestres', 'salas_ids'])
            
            # Datos
            for c in carreras:
                # Obtener semestres
                semestres = self.dao.obtener_semestres_carrera(c['id'])
                semestres_str = ','.join(map(str, semestres)) if semestres else ''
                
                # Obtener salas
                salas = self.dao.obtener_salas_carrera(c['id'])
                salas_ids_str = ','.join(str(s['id']) for s in salas) if salas else ''
                
                writer.writerow([
                    c['id'],
                    c['nombre'],
                    c.get('jornada', ''),
                    c.get('alumnos_proyectados', 0),
                    semestres_str,
                    salas_ids_str
                ])
        
        logging.info(f"✅ Exportadas {len(carreras)} carreras a {filepath}")
        return str(filepath)
    
    def exportar_modulos_csv(self, filepath: str = None) -> str:
        """Exporta todos los módulos a un archivo CSV"""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"modulos_{timestamp}.csv"
        
        modulos = self.dao.obtener_modulos()
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Encabezados
            writer.writerow(['id', 'nombre', 'codigo', 'horas_teoricas', 'horas_practicas', 
                           'alumnos_proyectados', 'semestre', 'carrera_id', 'docente_id', 'sala_id'])
            
            # Datos
            for m in modulos:
                writer.writerow([
                    m['id'],
                    m['nombre'],
                    m.get('codigo', ''),
                    m.get('horas_teoricas', 0),
                    m.get('horas_practicas', 0),
                    m.get('alumnos_proyectados', 0),
                    m.get('semestre', 0),
                    m.get('carrera_id', ''),
                    m.get('docente_id', ''),
                    m.get('sala_id', '')
                ])
        
        logging.info(f"✅ Exportados {len(modulos)} módulos a {filepath}")
        return str(filepath)
    
    def exportar_horarios_modulos_csv(self, filepath: str = None) -> str:
        """Exporta todos los horarios de módulos a un archivo CSV"""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"modulo_horarios_{timestamp}.csv"
        
        modulos = self.dao.obtener_modulos()
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Encabezados
            writer.writerow(['modulo_id', 'dia', 'hora_inicio', 'hora_fin'])
            
            # Datos
            total_horarios = 0
            for m in modulos:
                horarios = self.dao.obtener_horarios_modulo(m['id'])
                for h in horarios:
                    writer.writerow([
                        m['id'],
                        h['dia'],
                        h['hora_inicio'],
                        h['hora_fin']
                    ])
                    total_horarios += 1
        
        logging.info(f"✅ Exportados {total_horarios} horarios de módulos a {filepath}")
        return str(filepath)
    
    def exportar_disponibilidad_docentes_csv(self, filepath: str = None) -> str:
        """Exporta la disponibilidad de todos los docentes a un archivo CSV"""
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"disponibilidad_docentes_{timestamp}.csv"
        
        docentes = self.dao.obtener_docentes()
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            # Encabezados
            writer.writerow(['docente_id', 'dia', 'hora', 'estado'])
            
            # Datos
            total_bloques = 0
            for d in docentes:
                disponibilidad = self.dao.obtener_disponibilidad_docente(d['id'])
                for dia, horas in disponibilidad.items():
                    for hora, disponible in horas.items():
                        estado = 'disponible' if disponible else 'ocupado'
                        writer.writerow([
                            d['id'],
                            dia,
                            hora,
                            estado
                        ])
                        total_bloques += 1
        
        logging.info(f"✅ Exportados {total_bloques} bloques de disponibilidad a {filepath}")
        return str(filepath)
    
    def exportar_todo_csv(self) -> dict:
        """
        Exporta todas las tablas a archivos CSV individuales
        
        Returns:
            dict: Diccionario con los paths de todos los archivos generados
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_subdir = self.export_dir / f"export_{timestamp}"
        export_subdir.mkdir(exist_ok=True)
        
        archivos = {}
        
        logging.info("📦 Iniciando exportación completa a CSV...")
        
        archivos['docentes'] = self.exportar_docentes_csv(export_subdir / "docentes.csv")
        archivos['salas'] = self.exportar_salas_csv(export_subdir / "salas.csv")
        archivos['carreras'] = self.exportar_carreras_csv(export_subdir / "carreras.csv")
        archivos['modulos'] = self.exportar_modulos_csv(export_subdir / "modulos.csv")
        archivos['horarios'] = self.exportar_horarios_modulos_csv(export_subdir / "modulo_horarios.csv")
        archivos['disponibilidad'] = self.exportar_disponibilidad_docentes_csv(export_subdir / "disponibilidad_docentes.csv")
        
        logging.info(f"✅ Exportación completa finalizada en: {export_subdir}")
        
        return archivos
    
    def exportar_todo_excel(self, filepath: str = None) -> str:
        """
        Exporta todas las tablas a un único archivo Excel con múltiples hojas
        
        Returns:
            str: Path del archivo Excel generado
        """
        if not filepath:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = self.export_dir / f"datos_completos_{timestamp}.xlsx"
        
        logging.info("📊 Iniciando exportación a Excel...")
        
        # Crear workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remover hoja por defecto
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 1. Hoja de Docentes
        ws_docentes = wb.create_sheet("Docentes")
        ws_docentes.append(['ID', 'Nombre', 'Título', 'Contrato', 'Horas Contratadas', 'Email', 'Evaluación'])
        
        docentes = self.dao.obtener_docentes()
        for d in docentes:
            ws_docentes.append([
                d['id'], d['nombre'], d.get('titulo', ''), d.get('contrato', ''),
                d.get('horas_contratadas', 0), d.get('email', ''), d.get('evaluacion', 0.0)
            ])
        
        # Aplicar estilos a encabezado
        for cell in ws_docentes[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 2. Hoja de Salas
        ws_salas = wb.create_sheet("Salas")
        ws_salas.append(['ID', 'Nombre', 'Capacidad', 'Tipo'])
        
        salas = self.dao.obtener_salas()
        for s in salas:
            ws_salas.append([s['id'], s['nombre'], s.get('capacidad', 0), s.get('tipo', '')])
        
        for cell in ws_salas[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 3. Hoja de Carreras
        ws_carreras = wb.create_sheet("Carreras")
        ws_carreras.append(['ID', 'Nombre', 'Jornada', 'Alumnos Proyectados', 'Semestres', 'Salas IDs'])
        
        carreras = self.dao.obtener_carreras()
        for c in carreras:
            semestres = self.dao.obtener_semestres_carrera(c['id'])
            semestres_str = ','.join(map(str, semestres)) if semestres else ''
            
            salas_carrera = self.dao.obtener_salas_carrera(c['id'])
            salas_ids_str = ','.join(str(s['id']) for s in salas_carrera) if salas_carrera else ''
            
            ws_carreras.append([
                c['id'], c['nombre'], c.get('jornada', ''), 
                c.get('alumnos_proyectados', 0), semestres_str, salas_ids_str
            ])
        
        for cell in ws_carreras[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 4. Hoja de Módulos
        ws_modulos = wb.create_sheet("Módulos")
        ws_modulos.append(['ID', 'Nombre', 'Código', 'Horas Teóricas', 'Horas Prácticas', 
                          'Alumnos Proyectados', 'Semestre', 'Carrera ID', 'Docente ID', 'Sala ID'])
        
        modulos = self.dao.obtener_modulos()
        for m in modulos:
            ws_modulos.append([
                m['id'], m['nombre'], m.get('codigo', ''), m.get('horas_teoricas', 0),
                m.get('horas_practicas', 0), m.get('alumnos_proyectados', 0), m.get('semestre', 0),
                m.get('carrera_id', ''), m.get('docente_id', ''), m.get('sala_id', '')
            ])
        
        for cell in ws_modulos[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 5. Hoja de Horarios
        ws_horarios = wb.create_sheet("Horarios Módulos")
        ws_horarios.append(['Módulo ID', 'Día', 'Hora Inicio', 'Hora Fin'])
        
        for m in modulos:
            horarios = self.dao.obtener_horarios_modulo(m['id'])
            for h in horarios:
                ws_horarios.append([m['id'], h['dia'], h['hora_inicio'], h['hora_fin']])
        
        for cell in ws_horarios[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 6. Hoja de Disponibilidad (solo resumen para no hacer el archivo muy grande)
        ws_disp = wb.create_sheet("Disponibilidad Docentes")
        ws_disp.append(['Docente ID', 'Día', 'Hora', 'Estado'])
        
        # Solo exportar los primeros 1000 registros para no saturar
        count = 0
        for d in docentes:
            if count >= 1000:
                break
            disponibilidad = self.dao.obtener_disponibilidad_docente(d['id'])
            for dia, horas in disponibilidad.items():
                for hora, disponible in horas.items():
                    if count >= 1000:
                        break
                    estado = 'disponible' if disponible else 'ocupado'
                    ws_disp.append([d['id'], dia, hora, estado])
                    count += 1
        
        for cell in ws_disp[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Ajustar anchos de columna
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        wb.save(filepath)
        
        logging.info(f"✅ Exportación a Excel completada: {filepath}")
        logging.info(f"   - {len(docentes)} docentes")
        logging.info(f"   - {len(salas)} salas")
        logging.info(f"   - {len(carreras)} carreras")
        logging.info(f"   - {len(modulos)} módulos")
        
        return str(filepath)


# Función de utilidad para uso directo
def exportar_datos_completos(formato='excel'):
    """
    Exporta todos los datos de la base de datos
    
    Args:
        formato: 'excel' o 'csv'
    
    Returns:
        str o dict: Path del archivo Excel o diccionario de paths CSV
    """
    dao = SistemaDAO()
    exporter = DataExporter(dao)
    
    if formato.lower() == 'excel':
        return exporter.exportar_todo_excel()
    else:
        return exporter.exportar_todo_csv()


if __name__ == "__main__":
    # Prueba de exportación
    print("🚀 Iniciando exportación de datos...")
    
    # Exportar a Excel
    excel_file = exportar_datos_completos('excel')
    print(f"\n✅ Archivo Excel generado: {excel_file}")
    
    # Exportar a CSV
    csv_files = exportar_datos_completos('csv')
    print(f"\n✅ Archivos CSV generados:")
    for nombre, path in csv_files.items():
        print(f"   - {nombre}: {path}")
